#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def verificar_uso_periodo(data, data_inicio):
    try:
        if pd.isna(data):
            return False
        
        if not isinstance(data, pd.Timestamp):
            data = pd.to_datetime(data)
        
        return data >= data_inicio
    except:
        return False


def calcular_data_inicio(meses):
    hoje = datetime.now()
    data_limite = hoje - timedelta(days=meses * 30)
    data_inicio = datetime(data_limite.year, data_limite.month, 1)
    return data_inicio


def executar_marcar_usuarios(arquivo_entrada, meses_atras, callback=None):
    def log(msg):
        if callback:
            callback(msg)
        else:
            print(msg)
    
    if not os.path.exists(arquivo_entrada):
        msg = "Erro: Arquivo '%s' nao encontrado." % arquivo_entrada
        log(msg)
        return False, msg
    
    log("Lendo arquivo: %s" % arquivo_entrada)
    try:
        df = pd.read_excel(arquivo_entrada)
    except Exception as e:
        msg = "Erro ao ler arquivo: %s" % str(e)
        log(msg)
        return False, msg
    
    if 'USR_ID' not in df.columns:
        msg = "Erro: Coluna 'USR_ID' nao encontrada"
        log(msg)
        return False, msg
    
    if 'ULT_USO' not in df.columns:
        msg = "Erro: Coluna 'ULT_USO' nao encontrada"
        log(msg)
        return False, msg
    
    log("Total de linhas: %d" % len(df))
    
    data_inicio = calcular_data_inicio(meses_atras)
    log("Período analisado: a partir de %s" % data_inicio.strftime("%d de %B de %Y"))
    
    usuarios_com_uso = {}
    
    for idx, row in df.iterrows():
        usr_id = row['USR_ID']
        data_uso = row['ULT_USO']
        
        if usr_id not in usuarios_com_uso:
            usuarios_com_uso[usr_id] = []
        
        tem_uso = verificar_uso_periodo(data_uso, data_inicio)
        usuarios_com_uso[usr_id].append({
            'linha': idx,
            'tem_uso': tem_uso
        })
    
    usuarios_para_marcar = set()
    
    for usr_id, ocorrencias in usuarios_com_uso.items():
        if len(ocorrencias) == 2:
            if ocorrencias[0]['tem_uso'] and ocorrencias[1]['tem_uso']:
                usuarios_para_marcar.add(usr_id)
        elif len(ocorrencias) == 1:
            if ocorrencias[0]['tem_uso']:
                usuarios_para_marcar.add(usr_id)
    
    total_usuarios = len(usuarios_com_uso)
    usuarios_marcados = len(usuarios_para_marcar)
    
    log("Usuarios para marcar: %d" % usuarios_marcados)
    log("Total de usuarios: %d" % total_usuarios)
    
    nome_base = os.path.splitext(arquivo_entrada)[0]
    extensao = os.path.splitext(arquivo_entrada)[1]
    arquivo_saida = "%s_marcado%s" % (nome_base, extensao)
    
    if extensao.lower() == '.xls':
        arquivo_saida = "%s_marcado.xlsx" % nome_base
    
    log("Salvando arquivo: %s" % arquivo_saida)
    df.to_excel(arquivo_saida, index=False, engine='openpyxl')
    
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    
    cor_verde = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    
    linhas_marcadas = 0
    for idx, row in df.iterrows():
        usr_id = row['USR_ID']
        if usr_id in usuarios_para_marcar:
            linha_excel = idx + 2
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=linha_excel, column=col).fill = cor_verde
            linhas_marcadas += 1
    
    ultima_linha = len(df) + 2
    ws.cell(row=ultima_linha, column=1).value = "Total: %d" % usuarios_marcados
    
    for col in range(1, len(df.columns) + 1):
        ws.cell(row=ultima_linha, column=col).font = ws.cell(row=ultima_linha, column=col).font.copy()
    
    wb.save(arquivo_saida)
    
    log("\nSucesso!")
    log("- %d linhas marcadas em verde" % linhas_marcadas)
    log("- Total de usuarios: %d" % total_usuarios)
    log("- Arquivo salvo como: %s" % arquivo_saida)
    
    return True, "Processamento concluído com sucesso", arquivo_saida
