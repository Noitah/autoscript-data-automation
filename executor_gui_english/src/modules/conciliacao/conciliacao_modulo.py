#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo adaptado de conciliação para aceitar parâmetros externos
"""

import pandas as pd
import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from pathlib import Path


PDF_PATTERN = re.compile(r'(\d{2})\s+(\d{2})\s+([\d.,]+)(?:\s+(\d))?.*\.pdf$', re.IGNORECASE)
TOLERANCIA_DIFERENCA = 0.01


def extrair_dados_nome_pdf(nome_arquivo):
    """Extrai dados do nome do arquivo PDF"""
    try:
        match = PDF_PATTERN.match(nome_arquivo)
        if not match:
            return None
        
        mes = int(match.group(1))
        dia = int(match.group(2))
        valor_str = match.group(3).replace('.', '').replace(',', '.')
        valor = float(valor_str)
        versao = int(match.group(4)) if match.group(4) else 1
        
        if not (1 <= mes <= 12) or not (1 <= dia <= 31):
            return None
        
        chave_data = f"{dia:02d}/{mes:02d}"
        return mes, dia, valor, versao, chave_data
        
    except Exception:
        return None


def carregar_dados_excel(caminho, callback=None):
    """Carrega dados do Excel"""
    def log(msg):
        if callback:
            callback(msg)
        print(msg)
    
    log(f"Carregando dados do Excel: {caminho}")
    
    if not os.path.exists(caminho):
        return None, f"Arquivo Excel não encontrado: {caminho}"
    
    try:
        df = pd.read_excel(caminho, usecols=[0, 4], names=['Data', 'Valor'])
        registros_iniciais = len(df)
        
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce', dayfirst=True)
        df.dropna(subset=['Data'], inplace=True)
        
        registros_apos_data = len(df)
        if registros_iniciais != registros_apos_data:
            log(f"⚠️ {registros_iniciais - registros_apos_data} registros removidos por datas inválidas")
        
        df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
        df.fillna({'Valor': 0}, inplace=True)
        
        log(f"✅ {len(df)} registros válidos carregados do Excel")
        return df, None
        
    except Exception as e:
        return None, f"Falha ao ler o arquivo Excel: {e}"


def somar_valores_pdfs(pasta_pdf, callback=None):
    """Soma valores dos PDFs baseado nos nomes dos arquivos"""
    def log(msg):
        if callback:
            callback(msg)
        print(msg)
    
    log(f"Processando arquivos PDF na pasta: {pasta_pdf}")
    
    if not os.path.isdir(pasta_pdf):
        return None, 0, f"Pasta de PDFs não encontrada: {pasta_pdf}"
    
    todos_arquivos = [f for f in os.listdir(pasta_pdf) if f.lower().endswith('.pdf')]
    
    if not todos_arquivos:
        log("⚠️ Nenhum arquivo PDF encontrado na pasta")
        return pd.DataFrame(columns=['Data', 'Soma PDFs']), 0, None
    
    log(f"📄 Encontrados {len(todos_arquivos)} arquivos PDF")
    
    valores_pdfs = {}
    arquivos_ignorados = []
    
    for i, arquivo in enumerate(todos_arquivos, 1):
        if i % 50 == 0:
            log(f"Processando: {i}/{len(todos_arquivos)} arquivos...")
        
        dados = extrair_dados_nome_pdf(arquivo)
        if dados:
            mes, dia, valor, versao, chave_data = dados
            if versao == 1:
                valores_pdfs[chave_data] = valores_pdfs.get(chave_data, 0) + valor
        else:
            arquivos_ignorados.append(arquivo)
    
    if arquivos_ignorados:
        log(f"⚠️ {len(arquivos_ignorados)} arquivos PDF ignorados por não seguirem o padrão")
    
    df_pdf = pd.DataFrame(list(valores_pdfs.items()), columns=['Data', 'Soma PDFs'])
    log(f"✅ {len(df_pdf)} dias com valores de PDF consolidados")
    return df_pdf, len(arquivos_ignorados), None


def gerar_relatorio_conciliado(df_excel, df_pdfs, callback=None):
    """Gera relatório de conciliação"""
    def log(msg):
        if callback:
            callback(msg)
        print(msg)
    
    log("Gerando relatório de conciliação...")
    
    df_excel_agrupado = (
        df_excel.groupby(df_excel['Data'].dt.strftime('%d/%m'))['Valor']
        .sum()
        .reset_index()
    )
    df_excel_agrupado.rename(columns={'Valor': 'Valor Excel'}, inplace=True)

    df_final = pd.merge(df_excel_agrupado, df_pdfs, on='Data', how='outer')
    df_final.fillna(0, inplace=True)

    df_final['Diferenca'] = df_final['Valor Excel'] - df_final['Soma PDFs']

    df_final['Status'] = df_final['Diferenca'].apply(
        lambda x: 'OK' if abs(x) < TOLERANCIA_DIFERENCA else 'DIVERGE'
    )
    
    df_final['Saldo'] = df_final['Diferenca'].apply(
        lambda x: 'OK' if abs(x) < TOLERANCIA_DIFERENCA else 
                 ('FALTA VALOR NO PDF' if x > 0 else 'SOBRA VALOR NO PDF')
    )

    current_year = str(datetime.now().year)
    df_final['DataObj'] = pd.to_datetime(
        df_final['Data'] + '/' + current_year, 
        format='%d/%m/%Y'
    )
    df_final.sort_values('DataObj', inplace=True)
    df_final.drop(columns=['DataObj'], inplace=True)

    log(f"✅ Relatório gerado com {len(df_final)} registros")
    return df_final


def aplicar_formatacao_excel(ws):
    """Aplica formatação ao Excel"""
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    currency_format = 'R$ #,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    column_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 22}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in [2, 3, 4]:
                cell.number_format = currency_format

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    ws.conditional_formatting.add(
        f'E2:E{ws.max_row}',
        FormulaRule(formula=[f'E2="DIVERGE"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f'E2:E{ws.max_row}',
        FormulaRule(formula=[f'E2="OK"'], fill=green_fill)
    )

    ws.freeze_panes = 'A2'


def executar_conciliacao(caminho_excel, pasta_pdfs, pasta_saida=None, callback=None):
    """
    Executa o processo de conciliação completo
    
    Args:
        caminho_excel: Caminho para o arquivo Excel
        pasta_pdfs: Caminho para a pasta com PDFs
        pasta_saida: Pasta onde salvar o resultado (opcional)
        callback: Função para reportar progresso
    
    Returns:
        dict: (sucesso: bool, mensagem: str, caminho_arquivo: str)
    """
    def log(msg):
        if callback:
            callback(msg)
        print(msg)
    
    try:
        log('=== Iniciando processo de conciliação ===')
        
        # Definir pasta de saída
        if pasta_saida is None:
            home = Path.home()
            desktop_names = ['Desktop', 'Área de Trabalho', 'Escritorio']
            for name in desktop_names:
                desktop_path = home / name
                if desktop_path.exists():
                    pasta_saida = desktop_path / "Conciliação"
                    break
            if pasta_saida is None:
                pasta_saida = home / 'Desktop' / 'Conciliação'
        
        os.makedirs(pasta_saida, exist_ok=True)
        log(f'📁 Pasta de saída: {pasta_saida}')
        
        # Carregar dados do Excel
        df_excel, erro = carregar_dados_excel(caminho_excel, callback)
        if df_excel is None:
            return False, erro, None
        
        # Processar PDFs
        df_pdfs, arquivos_ignorados, erro = somar_valores_pdfs(pasta_pdfs, callback)
        if df_pdfs is None:
            return False, erro, None
        
        # Gerar relatório
        df_resultado = gerar_relatorio_conciliado(df_excel, df_pdfs, callback)
        
        # Salvar arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f'conciliacao_final_{timestamp}.xlsx'
        output_path = os.path.join(pasta_saida, output_filename)
        
        log(f"💾 Salvando relatório...")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_resultado.to_excel(writer, index=False, sheet_name='Conciliacao')
            ws = writer.sheets['Conciliacao']
            aplicar_formatacao_excel(ws)
        
        log('=== Processo concluído com sucesso ===')
        
        mensagem = f"""✅ Conciliação concluída!
📊 Estatísticas:
   - Registros Excel: {len(df_excel)}
   - Dias PDF: {len(df_pdfs)}
   - Registros relatório: {len(df_resultado)}
   - Arquivos ignorados: {arquivos_ignorados}
📁 Arquivo gerado: {output_path}"""
        
        return True, mensagem, output_path
        
    except Exception as e:
        return False, f"Erro crítico no processo: {e}", None


if __name__ == "__main__":
    print("Módulo de Conciliação - Teste")
